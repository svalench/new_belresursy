from openpyxl import load_workbook
from datetime import datetime
from num2t import num2text, decimal2text

# ----------------------------------------------------------------------------------------------------------------------
dt = datetime.now()
monat_to_ru = ["января", "февраля", "марта", "апреля", "мая", "июня", "июля", "августа",
               "сентября", "октября", "ноября", "декабря"]
# ----------------------------------------------------------------------------------------------------------------------
dataInvoice = {
    'date': '',                 # дата
    'auto': 'default auto',     # автомобиль
    'trailer': 'trailer',       # прицеп
    'waybill': 'waybill',       # путевой лист
    'auto_owner': 'auto_owner', # владелец автомобильной перевозки (плательщик)
    'driver': 'driver',         # водитель


    'customer_transportation':'',   # заказчик автомобильной перевозки
    'shipper':'',                   # грузоотправитель
    'consignee': '',                # грузополучатель
    'osnovanie_otpuska': '',        # основание отпуска
    'loading_point': '',            # пункт загрузки
    'unloading_point': '',          # пункт разгрузки

    # ТОВАРНЫЙ РАЗДЕЛ
    'value': '',        # количество
    'price': '',        # цена
    'netto': '',        # стоимость
    'vatSP': '20',      # ставка НДС
    'vat': 0,          # сумма НДС
    'brutto': 0,       # стоимость с НДС
    'value_place': '',  # колличество грузовых мест
    'weight': 0,        # массв груза

    # ПОГРУЗОЧНО-РАЗГРУЗОЧНЫЕ РАБОТЫ
    'time_arrival': '', # время прибытия
    'time_depart': '',  # время убытия
}
# ----------------------------------------------------------------------------------------------------------------------
def n2t(dec):
    return decimal2text( dec,
                        int_units=((u'рубль', u'рубля', u'рублей'), 'm'),
                        exp_units=((u'коп', u'коп', u'коп'), 'f'),
                        end_text='коп')
def n2tWeigth(dec):
    return decimal2text( dec,
                        int_units=((u'тн', u'тн', u'тн'), 'm'),
                        exp_units=((u'кг', u'кг', u'кг'), 'f'),
                        end_text='кг')
# ----------------------------------------------------------------------------------------------------------------------
# вертикальный формат авто
def invoiceBelcvetmet(path_to_SH, path_to_save, Data,response):
    # откроем шаблон
    xls = load_workbook(filename=path_to_SH)
    # получаем рабочий лист (первый)
    _xls = xls.active
    # _xls['I2'] = Data['UNP01']

    y = Data['date'][0:4]
    m = Data['date'][5:7]
    if m.isdigit():
        mm = monat_to_ru[int(m)-1]
    d = Data['date'][8:10]

    _xls['N2'] = Data['unpGruzopoluchatel']

    _xls['A16'] = d+' ' + mm+' '+y
    _xls['D17'] = Data['auto']
    _xls['P17'] = Data['trailer']
    _xls['Z17'] = Data['waybill']
    # _xls['E19'] = Data['auto_owner'] # остается неизменным
    _xls['X19'] = Data['driver']
    # _xls['J21'] = Data['customer_transportation']
    # _xls['E23'] = Data['shipper']
    _xls['E25'] = Data['consignee']
    _xls['C62'] = Data['consignee']
    _xls['E27'] = Data['osnovanie_otpuska']
    # _xls['M27'] = Data['loading_point']
    _xls['W27'] = Data['unloading_point']
    # ТОВАРНЫЙ РАЗДЕЛ
    _xls['A35'] = Data['name_tovar'] #наименование товара
    _xls['J35'] = Data['value']    # количество
    _xls['H35'] = Data['value_name']  # единица измерения
    _xls['L35'] = Data['price']        # цена
    _xls['N35'] = Data['netto']       # стоимость
    _xls['P35'] = Data['vatSP']      # ставка НДС
    _xls['R35'] = Data['vat']        # сумма НДС
    _xls['T35'] = Data['brutto']       # стоимость с НДС
    _xls['V35'] = Data['value_place']  # колличество грузовых мест
    _xls['X35'] = Data['weight']       # массв груза
    # цифры прописью
    _xls['E40'] = n2t(Data['brutto'])
    _xls['D42'] = n2t(Data['vat'])
    _xls['D44'] = n2tWeigth(Data['weight'])
    # погрузочно-разгрузочные работы
    _xls['I61'] = Data['time_arrival']      # время прибытия
    _xls['K61'] = Data['time_depart']

    xls.save(response)
def invoiceChermet(path_to_SH, path_to_save, Data,response):
    # откроем шаблон
    xls = load_workbook(filename=path_to_SH)
    # получаем рабочий лист (первый)
    _xls = xls.active
    _xls['A16'] = Data['date']
    _xls['D17'] = Data['auto']
    _xls['P17'] = Data['trailer']
    _xls['Z17'] = Data['waybill']
    # _xls['E19'] = Data['auto_owner'] # остается неизменным
    _xls['X19'] = Data['driver']
    # _xls['J21'] = Data['customer_transportation']
    # _xls['E23'] = Data['shipper']
    # _xls['E25'] = Data['consignee']
    _xls['E27'] = Data['osnovanie_otpuska']
    # _xls['M27'] = Data['loading_point']
    _xls['W27'] = Data['unloading_point']
    # ТОВАРНЫЙ РАЗДЕЛ
    _xls['J35'] = Data['value']    # количество
    _xls['L35'] = Data['price']        # цена
    _xls['N35'] = Data['netto']       # стоимость
    _xls['P35'] = Data['vatSP']      # ставка НДС
    _xls['R35'] = Data['vat']        # сумма НДС
    _xls['T35'] = Data['brutto']       # стоимость с НДС
    _xls['V35'] = Data['value_place']  # колличество грузовых мест
    _xls['X35'] = Data['weight']       # массв груза
    # цифры прописью
    _xls['E40'] = n2t(Data['brutto'])
    _xls['D42'] = n2t(Data['vat'])
    _xls['D43'] = n2tWeigth(Data['weight'])
    # погрузочно-разгрузочные работы
    _xls['I61'] = Data['time_arrival']      # время прибытия
    _xls['K61'] = Data['time_depart']

    xls.save(response)
def invoicePET(path_to_SH, path_to_save, Data,response):
    # откроем шаблон
    xls = load_workbook(filename=path_to_SH)
    # получаем рабочий лист (первый)
    _xls = xls.active
    _xls['A16'] = Data['date']
    _xls['D18'] = Data['auto']
    _xls['P18'] = Data['trailer']
    _xls['Z18'] = Data['waybill']
    # _xls['E20'] = Data['auto_owner'] # остается неизменным
    _xls['X20'] = Data['driver']
    # _xls['J21'] = Data['customer_transportation']
    # _xls['E23'] = Data['shipper']
    # _xls['E25'] = Data['consignee']
    _xls['E28'] = Data['osnovanie_otpuska']
    # _xls['M27'] = Data['loading_point']
    _xls['W28'] = Data['unloading_point']
    # ТОВАРНЫЙ РАЗДЕЛ
    _xls['J36'] = Data['value']    # количество
    _xls['L36'] = Data['price']        # цена
    _xls['N36'] = Data['netto']       # стоимость
    _xls['P36'] = Data['vatSP']      # ставка НДС
    _xls['R36'] = Data['vat']        # сумма НДС
    _xls['T36'] = Data['brutto']       # стоимость с НДС
    _xls['V36'] = Data['value_place']  # колличество грузовых мест
    _xls['X36'] = Data['weight']       # массв груза
    # цифры прописью
    _xls['E41'] = n2t(Data['brutto'])
    _xls['D43'] = n2t(Data['vat'])
    _xls['D44'] = n2tWeigth(Data['weight'])
    # погрузочно-разгрузочные работы
    _xls['I61'] = Data['time_arrival']      # время прибытия
    _xls['K61'] = Data['time_depart']

    xls.save(response)
# горизонтальный формат авто
def invoiceBelcvetmetH(path_to_SH, path_to_save, Data,response):
    # откроем шаблон
    xls = load_workbook(filename=path_to_SH)
    # sheets = xls.get_sheet_by_name('ТТН(книжн.) лиц.')
    # получаем рабочий лист (первый)
    _xls = xls.active
    sh02 = xls['ТТН(книжн.) лиц.']
    y = Data['date'][0:4]
    m = Data['date'][5:7]
    if m.isdigit():
        mm = monat_to_ru[int(m)-1]
    d = Data['date'][8:10]
    sh02['B19'] = d + ' ' + mm + ' ' + y
    #sh02['B19'] = Data['date'].day
    #sh02['F19'] = monat_to_ru[Data['date'].month-1]
    #sh02['X19'] = Data['date'].year - 2000

    sh02['N21'] = Data['auto']
    sh02['BK21'] = Data['trailer']
    sh02['V24'] = Data['waybill']
    # _xls['E19'] = Data['auto_owner'] # остается неизменным
    sh02['K26'] = Data['driver']
    # _xls['J21'] = Data['customer_transportation']
    # _xls['E23'] = Data['shipper']
    sh02['B51'] = Data['name_tovar']
    sh02['U37'] = Data['osnovanie_otpuska']
    # _xls['M27'] = Data['loading_point']
    sh02['R40'] = Data['unloading_point']
    # ТОВАРНЫЙ РАЗДЕЛ
    sh02['AD51'] = Data['value']    # количество
    sh02['AO51'] = Data['price']        # цена
    sh02['AW51'] = Data['netto']       # стоимость
    sh02['BG51'] = Data['vatSP']      # ставка НДС
    sh02['BO51'] = Data['vat']        # сумма НДС
    sh02['BW51'] = Data['brutto']       # стоимость с НДС
    sh02['CG51'] = Data['value_place']  # колличество грузовых мест
    sh02['CR51'] = Data['weight']       # массв груза
    # цифры прописью
    sh02['T54'] = n2t(Data['brutto'])
    sh02['Y56'] = n2t(Data['vat'])
    sh02['T59'] = n2tWeigth(Data['weight'])
    # # погрузочно-разгрузочные работы
    sh03 = xls['ТТН (книжн.) оборотн.']
    sh03['AI8'] = Data['time_arrival']      # время прибытия
    sh03['AP8'] = Data['time_depart']

    xls.save(response)
def invoiceChermetH(path_to_SH, path_to_save, Data,response):
    # откроем шаблон
    xls = load_workbook(filename=path_to_SH)
    # sheets = xls.get_sheet_by_name('ТТН(книжн.) лиц.')
    # получаем рабочий лист (первый)
    _xls = xls.active
    sh02 = xls['ТТН(книжн.) лиц.']
    sh02['B19'] = Data['date'].day
    sh02['F19'] = monat_to_ru[Data['date'].month-1]
    sh02['X19'] = Data['date'].year - 2000

    sh02['N21'] = Data['auto']
    sh02['BK21'] = Data['trailer']
    sh02['V24'] = Data['waybill']
    # _xls['E19'] = Data['auto_owner'] # остается неизменным
    sh02['K26'] = Data['driver']
    # _xls['J21'] = Data['customer_transportation']
    # _xls['E23'] = Data['shipper']
    # _xls['E25'] = Data['consignee']
    sh02['U37'] = Data['osnovanie_otpuska']
    # _xls['M27'] = Data['loading_point']
    sh02['R40'] = Data['unloading_point']
    # ТОВАРНЫЙ РАЗДЕЛ
    sh02['AD51'] = Data['value']    # количество
    sh02['AO51'] = Data['price']        # цена
    sh02['AW51'] = Data['netto']       # стоимость
    sh02['BG51'] = Data['vatSP']      # ставка НДС
    sh02['BO51'] = Data['vat']        # сумма НДС
    sh02['BW51'] = Data['brutto']       # стоимость с НДС
    sh02['CG51'] = Data['value_place']  # колличество грузовых мест
    sh02['CR51'] = Data['weight']       # массв груза
    # цифры прописью
    sh02['T54'] = n2t(Data['brutto'])
    sh02['Y56'] = n2t(Data['vat'])
    sh02['T59'] = n2tWeigth(Data['weight'])
    # # погрузочно-разгрузочные работы
    sh03 = xls['ТТН (книжн.) оборотн.']
    sh03['AI8'] = Data['time_arrival']      # время прибытия
    sh03['AP8'] = Data['time_depart']

    xls.save(response)
def invoicePETH(path_to_SH, path_to_save, Data,response):
    # откроем шаблон
    xls = load_workbook(filename=path_to_SH)
    # sheets = xls.get_sheet_by_name('ТТН(книжн.) лиц.')
    # получаем рабочий лист (первый)
    _xls = xls.active
    sh02 = xls['ТТН(книжн.) лиц.']
    sh02['B19'] = Data['date'].day
    sh02['F19'] = monat_to_ru[Data['date'].month-1]
    sh02['X19'] = Data['date'].year - 2000

    sh02['N21'] = Data['auto']
    sh02['BK21'] = Data['trailer']
    sh02['V24'] = Data['waybill']
    # _xls['E19'] = Data['auto_owner'] # остается неизменным
    sh02['K26'] = Data['driver']
    # _xls['J21'] = Data['customer_transportation']
    # _xls['E23'] = Data['shipper']
    # _xls['E25'] = Data['consignee']
    sh02['U37'] = Data['osnovanie_otpuska']
    # _xls['M27'] = Data['loading_point']
    sh02['R40'] = Data['unloading_point']
    # ТОВАРНЫЙ РАЗДЕЛ
    sh02['AD51'] = Data['value']    # количество
    sh02['AO51'] = Data['price']        # цена
    sh02['AW51'] = Data['netto']       # стоимость
    sh02['BG51'] = Data['vatSP']      # ставка НДС
    sh02['BO51'] = Data['vat']        # сумма НДС
    sh02['BW51'] = Data['brutto']       # стоимость с НДС
    sh02['CG51'] = Data['value_place']  # колличество грузовых мест
    sh02['CR51'] = Data['weight']       # массв груза
    # цифры прописью
    sh02['T54'] = n2t(Data['brutto'])
    sh02['Y56'] = n2t(Data['vat'])
    sh02['T59'] = n2tWeigth(Data['weight'])
    # # погрузочно-разгрузочные работы
    sh03 = xls['ТТН (книжн.) оборотн.']
    sh03['AI8'] = Data['time_arrival']      # время прибытия
    sh03['AP8'] = Data['time_depart']

    xls.save(response)
# железнодорожные вагоны
def invoiceZDotsev(path_to_SH, path_to_save, Data):
    # откроем шаблон
    xls = load_workbook(filename=path_to_SH)
    # получаем рабочий лист (первый)
    _xls = xls.active
    _xls['C11'] = Data['date']
    # _xls['E13'] = Data['customer_transportation']
    # _xls['E13'] = "Data['shipper']"
    # _xls['E15'] = "Data['consignee']"
    _xls['E19'] = "Data['osnovanie_otpuska']"
    # ТОВАРНЫЙ РАЗДЕЛ
    _xls['H26'] = Data['value']    # количество
    _xls['N26'] = Data['price']        # цена
    _xls['P26'] = Data['netto']       # стоимость
    _xls['W26'] = Data['vatSP']      # ставка НДС
    _xls['X26'] = Data['vat']        # сумма НДС
    _xls['Y26'] = Data['brutto']       # стоимость с НДС
    # _xls['V36'] = Data['value_place']  # колличество грузовых мест
    # _xls['X36'] = Data['weight']       # массв груза
    # цифры прописью
    _xls['E32'] = n2t(Data['brutto'])
    _xls['E30'] = n2t(Data['vat'])
    # _xls['D44'] = n2tWeigth(Data['weight'])
    # погрузочно-разгрузочные работы
    # _xls['I61'] = Data['time_arrival']      # время прибытия
    # _xls['K61'] = Data['time_depart']

    xls.save(response)
def invoiceZDstekloboiWHITE(path_to_SH, path_to_save, Data):
    # откроем шаблон
    xls = load_workbook(filename=path_to_SH)
    # получаем рабочий лист (первый)
    _xls = xls.active
    _xls['D12'] = Data['date']
    _xls['E18'] = "Data['osnovanie_otpuska']"
    # ТОВАРНЫЙ РАЗДЕЛ
    _xls['H25'] = Data['value']    # количество
    _xls['N25'] = Data['price']        # цена
    _xls['P25'] = Data['netto']       # стоимость
    _xls['W25'] = Data['vatSP']      # ставка НДС
    _xls['X25'] = Data['vat']        # сумма НДС
    _xls['Y25'] = Data['brutto']       # стоимость с НДС
    # цифры прописью
    _xls['E28'] = n2t(Data['brutto'])
    _xls['E30'] = n2t(Data['vat'])
    # _xls['D44'] = n2tWeigth(Data['weight'])
    # погрузочно-разгрузочные работы
    # _xls['I61'] = Data['time_arrival']      # время прибытия
    # _xls['K61'] = Data['time_depart']

    xls.save(response)
def invoiceZDstekloboiGREEN(path_to_SH, path_to_save, Data):
    # откроем шаблон
    xls = load_workbook(filename=path_to_SH)
    # получаем рабочий лист (первый)
    _xls = xls.active
    _xls['C11'] = Data['date']
    _xls['E17'] = "Data['osnovanie_otpuska']"
    # ТОВАРНЫЙ РАЗДЕЛ
    _xls['H24'] = Data['value']    # количество
    _xls['N24'] = Data['price']        # цена
    _xls['P24'] = Data['netto']       # стоимость
    _xls['W24'] = Data['vatSP']      # ставка НДС
    _xls['X24'] = Data['vat']        # сумма НДС
    _xls['Y24'] = Data['brutto']       # стоимость с НДС
    # цифры прописью
    _xls['E28'] = n2t(Data['brutto'])
    _xls['E30'] = n2t(Data['vat'])
    # _xls['D44'] = n2tWeigth(Data['weight'])
    # погрузочно-разгрузочные работы
    # _xls['I61'] = Data['time_arrival']      # время прибытия
    # _xls['K61'] = Data['time_depart']

    xls.save(response)
def invoiceZDstekloboiBROWN(path_to_SH, path_to_save, Data):
    # откроем шаблон
    xls = load_workbook(filename=path_to_SH)
    # получаем рабочий лист (первый)
    _xls = xls.active
    _xls['C11'] = Data['date']
    _xls['E17'] = "Data['osnovanie_otpuska']"
    # ТОВАРНЫЙ РАЗДЕЛ
    _xls['H24'] = Data['value']    # количество
    _xls['N24'] = Data['price']        # цена
    _xls['P24'] = Data['netto']       # стоимость
    _xls['W24'] = Data['vatSP']      # ставка НДС
    _xls['X24'] = Data['vat']        # сумма НДС
    _xls['Y24'] = Data['brutto']       # стоимость с НДС
    # цифры прописью
    _xls['E28'] = n2t(Data['brutto'])
    _xls['E30'] = n2t(Data['vat'])
    # _xls['D44'] = n2tWeigth(Data['weight'])
    # погрузочно-разгрузочные работы
    # _xls['I61'] = Data['time_arrival']      # время прибытия
    # _xls['K61'] = Data['time_depart']

    xls.save(response)
# ----------------------------------------------------------------------------------------------------------------------

if __name__ == '__main__':
    path00 = '/home/lex/PycharmProjects/new_bel/templates/doc/auto/invoice_horizontal/cvetmet.xlsx'
    path01 = '/home/lex/PycharmProjects/new_bel/templates/doc/auto/invoice_horizontal/chermet.xlsx'
    path02 = '/home/lex/PycharmProjects/new_bel/templates/doc/auto/invoice_horizontal/pet.xlsx'
    path000= 'C:/BELRESURS/python/print_form/invoice_template/auto/invoice_vertical/Накл Цветмет верт.xlsx'
    path001 = 'C:/BELRESURS/python/print_form/invoice_template/auto/invoice_vertical/Накл Чермет верт.xlsx'
    path002 = 'C:/BELRESURS/python/print_form/invoice_template/auto/invoice_vertical/Накл ПЭТ верт.xlsx'
    path0001 = 'C:/BELRESURS/python/print_form/invoice_template/wagon/Товарная накладная отсев(продукция).xlsx'
    path0002 = 'C:/BELRESURS/python/print_form/invoice_template/wagon/Товарная накладная стеклобой бесцветный.xlsx'
    path0003 = 'C:/BELRESURS/python/print_form/invoice_template/wagon/Товарная накладная стеклобой зеленый.xlsx'
    path0004 = 'C:/BELRESURS/python/print_form/invoice_template/wagon/Товарная накладная стеклобой коричневый.xlsx'

    pathS = 'result001.xlsx'
    dataInvoice['time_arrival'] = dt
    dataInvoice['time_depart'] = dt
    dataInvoice['date'] = dt

    # invoiceBelcvetmet(path00, pathS, Data=dataInvoice)
    # invoiceChermet(path01, pathS, dataInvoice)
    # invoicePET(path02, pathS, dataInvoice)
    # invoiceBelcvetmetH(path000, pathS, Data=dataInvoice)
    # invoiceChermetH(path001, pathS, Data=dataInvoice)
    # invoicePETH(path002, pathS, Data=dataInvoice)
    # invoiceZDotsev(path0001, pathS, Data=dataInvoice)
    # invoiceZDstekloboiWHITE(path0002, pathS, Data=dataInvoice)
    # invoiceZDstekloboiGREEN(path0003, pathS, Data=dataInvoice)
    invoiceZDstekloboiBROWN(path0004, pathS, Data=dataInvoice)

